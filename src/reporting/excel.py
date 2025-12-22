"""
Excel Report Generator
=====================

Excel export using openpyxl.

Includes:
- Summary sheet (counts, statistics)
- Findings sheet (all findings in table format)
- Requirements sheet (all requirements checked)
- Filterable columns
- Conditional formatting for severity
"""

from typing import List, Dict
from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")

from .base import BaseReportGenerator


class ExcelReportGenerator(BaseReportGenerator):
    """
    Generate Excel compliance reports.
    """

    def __init__(self):
        """Initialize Excel generator."""
        super().__init__()

        # Define color schemes
        self.severity_colors = {
            'critical': 'FFDC3545',  # Red
            'high': 'FFFD7E14',      # Orange
            'medium': 'FFFFC107',    # Yellow
            'low': '0028A745',       # Green
        }

    def generate(self, findings: List[Dict], config: Dict) -> bytes:
        """
        Generate Excel report.

        Args:
            findings: List of findings
            config: Configuration dictionary

        Returns:
            bytes: Excel file as bytes
        """
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets
        self._create_summary_sheet(wb, findings, config)
        self._create_findings_sheet(wb, findings)
        self._create_statistics_sheet(wb, findings)

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    def _create_summary_sheet(self, wb: Workbook, findings: List[Dict], config: Dict):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = 'Compliance Gap Analysis Report'
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:D1')

        # Metadata
        ws['A3'] = 'Report Date:'
        ws['B3'] = self.format_date()
        ws['A4'] = 'Jurisdiction:'
        ws['B4'] = config.get('jurisdiction', 'N/A')
        ws['A5'] = 'Domains:'
        ws['B5'] = ', '.join(config.get('selected_domains', []))

        # Summary statistics
        summary = self.format_finding_summary(findings)

        ws['A7'] = 'Summary Statistics'
        ws['A7'].font = Font(size=14, bold=True)

        # Create summary table
        headers = ['Metric', 'Value']
        ws.append([])  # Empty row
        ws.append(headers)

        # Header styling
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF2E2E38', end_color='FF2E2E38', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        data = [
            ['Total Findings', summary['total_findings']],
            ['Critical Priority', summary['by_severity']['critical']],
            ['High Priority', summary['by_severity']['high']],
            ['Medium Priority', summary['by_severity']['medium']],
            ['Low Priority', summary['by_severity']['low']],
            ['Compliance Rate', f"{summary['compliance_rate']:.1f}%"],
        ]

        for row_data in data:
            ws.append(row_data)

        # Domain breakdown
        ws.append([])
        ws.append(['Findings by Domain'])
        ws[f'A{ws.max_row}'].font = Font(size=12, bold=True)

        ws.append(['Domain', 'Count'])
        header_row = ws.max_row
        for col in range(1, 3):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF2E2E38', end_color='FF2E2E38', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for domain, count in summary['by_domain'].items():
            ws.append([domain, count])

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_findings_sheet(self, wb: Workbook, findings: List[Dict]):
        """Create findings sheet with all findings."""
        ws = wb.create_sheet("Findings")

        # Headers
        headers = [
            'Severity',
            'Domain',
            'Requirement Citation',
            'Requirement Text',
            'Policy Reference',
            'Gap Description',
            'Recommendation',
            'Confidence'
        ]

        ws.append(headers)

        # Header styling
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF2E2E38', end_color='FF2E2E38', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Data rows
        for finding in findings:
            row_data = [
                finding.get('severity', 'medium').upper(),
                finding.get('domain', 'Unknown'),
                finding.get('requirement_citation', 'N/A'),
                finding.get('requirement_text', ''),
                finding.get('policy_reference', 'N/A'),
                finding.get('gap_description', ''),
                finding.get('recommendation', ''),
                self.format_confidence(finding.get('confidence', 0.0))
            ]

            ws.append(row_data)

            # Apply severity color coding
            row_num = ws.max_row
            severity = finding.get('severity', 'medium').lower()
            severity_color = self.severity_colors.get(severity, 'FFFFFFFF')

            # Color the severity cell
            severity_cell = ws.cell(row=row_num, column=1)
            severity_cell.fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type='solid')
            severity_cell.font = Font(bold=True, color='FFFFFFFF')
            severity_cell.alignment = Alignment(horizontal='center', vertical='top')

            # Wrap text for long columns
            for col in [3, 4, 5, 6, 7]:  # Text columns
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Adjust column widths
        column_widths = [12, 15, 25, 40, 20, 50, 50, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add autofilter
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'

    def _create_statistics_sheet(self, wb: Workbook, findings: List[Dict]):
        """Create detailed statistics sheet."""
        ws = wb.create_sheet("Statistics")

        # Title
        ws['A1'] = 'Detailed Statistics'
        ws['A1'].font = Font(size=16, bold=True)

        ws.append([])

        # Severity breakdown
        ws.append(['Severity Breakdown'])
        ws[f'A{ws.max_row}'].font = Font(size=12, bold=True)

        categorized = self.categorize_by_severity(findings)

        severity_data = [['Severity', 'Count', 'Percentage']]
        total = len(findings)

        for severity in ['critical', 'high', 'medium', 'low']:
            count = len(categorized.get(severity, []))
            percentage = (count / total * 100) if total > 0 else 0
            severity_data.append([severity.upper(), count, f"{percentage:.1f}%"])

        for row_data in severity_data:
            ws.append(row_data)

            if row_data[0] != 'Severity':  # Not header
                row_num = ws.max_row
                severity_lower = row_data[0].lower()
                severity_color = self.severity_colors.get(severity_lower, 'FFFFFFFF')

                # Color the severity cell
                severity_cell = ws.cell(row=row_num, column=1)
                severity_cell.fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type='solid')
                severity_cell.font = Font(bold=True, color='FFFFFFFF')

        ws.append([])

        # Domain breakdown
        ws.append(['Domain Breakdown'])
        ws[f'A{ws.max_row}'].font = Font(size=12, bold=True)

        categorized_domain = self.categorize_by_domain(findings)

        domain_data = [['Domain', 'Count', 'Percentage']]

        for domain, domain_findings in categorized_domain.items():
            count = len(domain_findings)
            percentage = (count / total * 100) if total > 0 else 0
            domain_data.append([domain, count, f"{percentage:.1f}%"])

        for row_data in domain_data:
            ws.append(row_data)

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
